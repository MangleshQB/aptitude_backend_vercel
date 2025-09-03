import pytz
import io

from django.db.models import Count
from django.utils import timezone
from datetime import datetime, timedelta
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from tracking.models import EmployeeTimeSheet
from app.models import CustomUser, PersonnelEmployee, IclockTransaction
import pandas as pd
from django.http import HttpResponse
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter
from utils.common import get_all_reporting_users, get_crm_leave, get_holidays


def format_seconds_to_hms(total_seconds):
    hours = int(total_seconds // 3600)
    minutes = int((total_seconds % 3600) // 60)
    seconds = int(total_seconds % 60)

    return f"{hours:02}:{minutes:02}:{seconds:02}"

def get_total_sum(datalist):
    final_total_working_hours = sum(
        timedelta(
            hours=int(day['total_working_hours'].split(':')[0]),
            minutes=int(day['total_working_hours'].split(':')[1]),
            seconds=int(day['total_working_hours'].split(':')[2])
        ).total_seconds()
        for day in datalist if day['total_working_hours']
    )

    final_total_effective_hours = sum(
        timedelta(
            hours=int(day['total_effective_hours'].split(':')[0]),
            minutes=int(day['total_effective_hours'].split(':')[1]),
            seconds=int(day['total_effective_hours'].split(':')[2])
        ).total_seconds()
        for day in datalist if day['total_effective_hours']
    )

    final_total_overtime = sum(
        timedelta(
            hours=int(day['total_over_time'].split(':')[0]),
            minutes=int(day['total_over_time'].split(':')[1]),
            seconds=int(day['total_over_time'].split(':')[2])
        ).total_seconds()
        for day in datalist if day['total_over_time']
    )

    final_total_break_time = sum(
        timedelta(
            hours=int(day['total_break_time'].split(':')[0]),
            minutes=int(day['total_break_time'].split(':')[1]),
            seconds=int(day['total_break_time'].split(':')[2])
        ).total_seconds()
        for day in datalist if day['total_break_time']
    )

    # final_total_idle_time = str(
    #     round(sum(float(sit['total_idle_time']) for sit in datalist if sit['total_idle_time']), 2))

    final_total_idle_time = sum(
        timedelta(
            hours=int(day['total_idle_time'].split(':')[0]),
            minutes=int(day['total_idle_time'].split(':')[1]),
            seconds=int(day['total_idle_time'].split(':')[2])
        ).total_seconds()
        for day in datalist if day['total_idle_time']
    )

    final_total_approved_idle_time = sum(
        timedelta(
            hours=int(day['total_approved_idle_time'].split(':')[0]),
            minutes=int(day['total_approved_idle_time'].split(':')[1]),
            seconds=int(day['total_approved_idle_time'].split(':')[2])
        ).total_seconds()
        for day in datalist if day['total_approved_idle_time']
    )

    return (format_seconds_to_hms(final_total_working_hours), format_seconds_to_hms(final_total_effective_hours),
            format_seconds_to_hms(final_total_overtime), format_seconds_to_hms(final_total_break_time),
            format_seconds_to_hms(final_total_idle_time), format_seconds_to_hms(final_total_approved_idle_time))


class ExportExcel(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        try:
            start_date_ = request.data.get('start_date')
            end_date_ = request.data.get('end_date')
            email_list = request.data.get('emails')

            if not isinstance(email_list, list):
                raise ValueError('emails must be a list')

            try:
                start_date = timezone.make_aware(datetime.strptime(start_date_, "%d/%m/%Y"), timezone=pytz.UTC)
                end_date = timezone.make_aware(datetime.strptime(end_date_, "%d/%m/%Y"), timezone=pytz.UTC)
            except:
                raise ValueError('wrong date format, valid format is DD/MM/YYYY')

            permissions = (request.user.groups.all().first().permissions.filter(content_type__model='customuser')
                           .values_list('codename', flat=True))
            base_permissions = [permission.split('_')[0] for permission in permissions]
            # team = request.user.subordinates.values_list('email', flat=True)
            reporting_user = get_all_reporting_users(request.user)
            team = [i.email for i in reporting_user]
            req_email = request.user.email
            final_df = []
            for email in email_list:
                user = ''
                if 'all' in base_permissions:
                    user = CustomUser.objects.filter(email=email).first()
                elif 'team' in base_permissions:
                    if email in team:
                        user = CustomUser.objects.filter(email=email).first()
                elif 'owned' in base_permissions:
                    if req_email == email:
                        user = CustomUser.objects.filter(email=email).first()
                if not user: raise ValueError('You do not have access!')

                data = EmployeeTimeSheet.objects.filter(
                    user__email=email,
                    date__range=[start_date, end_date]
                ).values()

                if not data:
                    continue
                data = list(data)
                (final_total_working_hours, final_total_effective_hours, final_total_overtime, final_total_break_time,
                 final_total_idle_time, final_total_approved_idle_time) = get_total_sum(data)
                data.append({
                    'date': '',
                    'name': '',
                    'employee_id': 'Total',
                    'total_effective_hours': final_total_effective_hours,
                    'total_working_hours': final_total_working_hours,
                    'total_over_time': final_total_overtime,
                    'total_break_time': final_total_break_time,
                    'total_idle_time': final_total_idle_time,
                    'total_approved_idle_time': final_total_approved_idle_time,
                    'first_punch_in': '',
                    'last_punch_out': '',
                    'is_late_coming': '',
                    'missing_punch_in': '',
                    'missing_punch_out': '',
                    'status': '',
                })


                df = pd.DataFrame(data)

                # df['name'] = f'{str(user.first_name) if user.first_name else ""} {str(user.last_name) if user.last_name else ""}'.strip().title()
                # df['employee_id'] = user.employee_id

                df['name'] = df.apply(
                    lambda
                        row: f"{str(user.first_name) if user.first_name else ''} {str(user.last_name) if user.last_name else ''}".strip().title()
                    if row['employee_id'] != 'Total' else '',
                    axis=1
                )

                df['employee_id'] = df.apply(
                    lambda row: user.employee_id if str(row['employee_id']) != 'Total' else 'Total',
                    axis=1
                )

                columns_to_include = [
                    'date',
                    'name',
                    'employee_id',
                    'total_effective_hours',
                    'total_working_hours',
                    'total_over_time',
                    'total_break_time',
                    'total_idle_time',
                    'total_approved_idle_time',
                    'first_punch_in',
                    'last_punch_out',
                    'is_late_coming',
                    'missing_punch_in',
                    'missing_punch_out',
                    'status',
                ]
                df = df[columns_to_include]
                df.columns = [col.replace('_', ' ').title() for col in df.columns]

                for column in df.columns:
                    if df[column].dtype == 'datetime64[ns, UTC]':
                        df[column] = df[column].dt.tz_localize(None)
                        df[column] = df[column].dt.strftime('%d-%m-%Y %H:%M:%S')
                    elif df[column].dtype == 'datetime64[ns]':
                        df[column] = df[column].dt.strftime('%d-%m-%Y %H:%M:%S')
                    elif df[column].dtype == 'bool':
                        df[column] = df[column].apply(lambda x: 'Yes' if x else 'No')
                    elif df[column].dtype == 'object':
                        df[column] = df[column].fillna('')

                final_df.append(df)

            if final_df:
                final_df = sorted(final_df, key=lambda df: df['Name'].iloc[0].split()[0].lower())

                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    header_fill = PatternFill(start_color="83a0de", end_color="83a0de", fill_type="solid")
                    sheet_names = []
                    for df in final_df:
                        sheet_name = df['Name'].iloc[0].title()
                        sheet_names.append(sheet_name)
                        df.to_excel(writer, index=False, sheet_name=sheet_name)

                        workbook = writer.book
                        worksheet = workbook[sheet_name]

                        # Apply alignment and border styles to all cells
                        for row in worksheet.iter_rows():
                            for cell in row:
                                cell.alignment = Alignment(horizontal='center', vertical='center')

                        border_style = Border(
                            left=Side(border_style='thin'),
                            right=Side(border_style='thin'),
                            top=Side(border_style='thin'),
                            bottom=Side(border_style='thin')
                        )
                        for row in worksheet.iter_rows():
                            for cell in row:
                                cell.border = border_style

                        # Apply header background color
                        for cell in worksheet[1]:
                            cell.fill = header_fill

                        # Apply column width adjustment
                        for col in worksheet.columns:
                            max_length = 0
                            column = col[0].column_letter

                            max_length = len(str(col[0].value))

                            for cell in col:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass

                                adjusted_width = max_length + 2
                                worksheet.column_dimensions[column].width = adjusted_width

                        # Add bold text formatting and retain background color for the last row
                        last_row_idx = worksheet.max_row  # Identify the last row index
                        last_row_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                                                    fill_type="solid")  # Light red fill
                        for cell in worksheet[last_row_idx]:
                            cell.font = Font(bold=True)  # Make text bold
                            if not cell.fill or cell.fill == PatternFill():  # Retain or apply background color
                                cell.fill = last_row_fill

                buffer.seek(0)
                response = HttpResponse(
                    buffer,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="{start_date_} To {end_date_}.xlsx"'
                return response
            else:
                raise ValueError('data not found!')

        except Exception as e:
            error = str(e)
            ret_status = status.HTTP_400_BAD_REQUEST

        return Response({'error': error}, status=ret_status)


def get_report(datalist):
    final_total_working_hours = sum(
        timedelta(
            hours=int(day['total_working_hours'].split(':')[0]),
            minutes=int(day['total_working_hours'].split(':')[1]),
            seconds=int(day['total_working_hours'].split(':')[2])
        ).total_seconds()
        for day in datalist if day['total_working_hours']
    )

    final_total_effective_hours = sum(
        timedelta(
            hours=int(day['total_effective_hours'].split(':')[0]),
            minutes=int(day['total_effective_hours'].split(':')[1]),
            seconds=int(day['total_effective_hours'].split(':')[2])
        ).total_seconds()
        for day in datalist if day['total_effective_hours']
    )

    total_late_coming = len([day['is_late_coming'] for day in datalist if day['is_late_coming']])

    return (final_total_working_hours, format_seconds_to_hms(final_total_working_hours),
            final_total_effective_hours, format_seconds_to_hms(final_total_effective_hours), total_late_coming)

class ExportReport(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):

        start_date_ = request.data.get('start_date')
        end_date_ = request.data.get('end_date')

        try:
            start_date = timezone.make_aware(datetime.strptime(start_date_, "%d/%m/%Y"), timezone=pytz.UTC)
            end_date = timezone.make_aware(datetime.strptime(end_date_, "%d/%m/%Y"), timezone=pytz.UTC)
        except:
            raise ValueError('wrong date format, valid format is DD/MM/YYYY')

        permissions = (request.user.groups.all().first().permissions.filter(content_type__model='customuser')
                       .values_list('codename', flat=True))
        base_permissions = [permission.split('_')[0] for permission in permissions]

        if 'all' in base_permissions:
            queryset = CustomUser.objects.all()
        elif 'team' in base_permissions:
            queryset = get_all_reporting_users(request.user)
        elif 'owned' in base_permissions:
            queryset = CustomUser.objects.filter(id=request.user.id).first()
        else:
            raise ValueError('')

        # all_holidays = get_holidays(start_date=start_date, end_date=end_date)

        final_dict = []
        for user in queryset:

            data = list(EmployeeTimeSheet.objects.filter(
                user__email=user.email,
                date__range=[start_date, end_date]
            ).values())

            if not data:continue

            (final_total_working_hours_second, final_total_working_hours, final_total_effective_hours_second,
             final_total_effective_hours, total_late_coming) = get_report(data)

            half_day, leave_taken, paid_leave, lwp, SL, CL = get_crm_leave(email=user.email, start_date=start_date, end_date=end_date)

            temp_dict = {}

            present_day = float(len(data)) - half_day

            over_time = final_total_effective_hours_second - (present_day * 8.5 * 3600)
            if over_time > 0.0:
                OT = format_seconds_to_hms(over_time)
            else:
                OT = ''

            total_working_days = len(pd.date_range(start=start_date, end=end_date, freq='B'))
            leave_without_approved = total_working_days - leave_taken - present_day
            if leave_without_approved <= 0.0:
                leave_without_approved = '0.0'

            temp_dict['name'] = f'{user.first_name} {user.last_name if user.last_name else ""}'
            temp_dict['employee_id'] = str(user.employee_id)
            temp_dict['total_working_days'] = str(total_working_days)
            temp_dict['leave_taken'] = str(leave_taken)
            temp_dict['paid_leave'] = str(paid_leave)
            temp_dict['lwp'] = str(lwp)
            temp_dict['SL'] = str(SL)
            temp_dict['CL'] = str(CL)
            temp_dict['present_day'] = str(present_day)
            temp_dict['leave_without_approved'] = str(leave_without_approved)
            temp_dict['actual_hours_need'] = format_seconds_to_hms(present_day * 8.5 * 3600)
            temp_dict['total_working_hours'] = final_total_working_hours
            temp_dict['total_effective_hours'] = final_total_effective_hours
            temp_dict['total_over_time'] = OT
            temp_dict['total_late_coming'] = str(total_late_coming)

            final_dict.append(temp_dict)

        final_df = pd.DataFrame(final_dict)
        final_df.columns = [col.replace('_', ' ').title() for col in final_df.columns]
        # Export to Excel
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # New Colors
            date_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Dark blue

            sheet_name = "Employee Report"
            final_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=1)  # Start from row 2

            workbook = writer.book
            worksheet = workbook[sheet_name]

            # Add the date range to the first row and merge cells
            date_range = f"Report: {start_date_} to {end_date_}"
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(final_df.columns))
            worksheet.cell(row=1, column=1).value = date_range
            worksheet.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
            worksheet.cell(row=1, column=1).fill = date_fill
            worksheet.cell(row=1, column=1).font = Font(bold=True, size=12)

            # Apply alignment and border styles to all data rows
            for row in worksheet.iter_rows(min_row=2):  # Start from the header row
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            border_style = Border(
                left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin')
            )
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.border = border_style

            # Apply header background color
            for cell in worksheet[2]:  # Header row is now the second row
                cell.fill = header_fill

            # Adjust column widths
            for col_idx, column_cells in enumerate(worksheet.columns, 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                for cell in column_cells:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                worksheet.column_dimensions[column_letter].width = max_length + 2

        buffer.seek(0)
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Employee Report_{start_date_} To {end_date_}.xlsx"'
        return response

